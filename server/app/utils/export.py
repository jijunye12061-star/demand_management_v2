import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# 管理端全量导出列
FULL_COLUMNS = [
    ("ID", "id", 8),
    ("标题", "title", 30),
    ("需求描述", "description", 40),
    ("需求类型", "request_type", 14),
    ("研究范畴", "research_scope", 12),
    ("机构", "org_name", 18),
    ("客户类型", "org_type", 10),
    ("部门", "department", 10),
    ("销售", "sales_name", 12),
    ("研究员", "researcher_name", 12),
    ("状态", "status", 10),
    ("工时(h)", "work_hours", 10),
    ("自动化建设工时(h)", "automation_hours", 14),
    ("总工时(h)", "total_work_hours", 10),
    ("关联需求ID", "parent_request_id", 12),
    ("创建时间", "created_at", 18),
    ("完成时间", "completed_at", 18),
    ("下载次数", "download_count", 10),
]

# FIX: 需求动态导出列 — api.md §8 明确要求 completed_at 而非 created_at
FEED_COLUMNS = [
    ("需求标题", "title", 30),
    ("需求描述", "description", 40),
    ("机构类型", "org_type", 10),
    ("需求类型", "request_type", 14),
    ("研究范围", "research_scope", 12),
    ("对接研究员", "researcher_name", 12),
    ("完成时间", "completed_at", 18),
]

# FIX: 补全 withdrawn / canceled, admin 全量导出会用到
STATUS_MAP = {
    "pending": "待处理",
    "in_progress": "处理中",
    "completed": "已完成",
    "withdrawn": "已退回",
    "canceled": "已取消",
}


def generate_excel(items: list[dict], columns: list[tuple] | None = None) -> io.BytesIO:
    cols = columns or FULL_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = "需求数据"

    # Header
    for col_idx, (label, _, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, (_, key, _) in enumerate(cols, 1):
            value = item.get(key, "")
            if key == "status":
                value = STATUS_MAP.get(value, value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
